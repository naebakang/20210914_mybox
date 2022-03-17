# File encoding: utf8

import os
# import win32gui
# import win32con
import time
import numpy
# import pyautogui
import openpyxl
from openpyxl.utils import get_column_letter
import pandas
from sympy import Symbol, solve
import json
import datetime

"""
작성규칙
내가 만든 모듈은 import 하지 않는다.

# symbol = 'ETHUSDT'
# df['open_time'] = pandas.to_datetime(df['open_time'], unit='ms')
# df.set_index('open_time', inplace=True)
# print(df)
"""

# class MSWindow:
#     def __init__(self):
#         pass
#
#     @staticmethod
#     def get_recent_file_name(directory, file_group_name, len_now_time):
#         """
#         :param directory:
#         :param file_group_name:
#         :param len_now_time:
#         :return: 최신 파일이름[str]
#         """
#
#         # parameter
#         directory = str(directory)
#         file_group_name = str(file_group_name)
#         len_now_time = int(len_now_time)
#
#         file_name_list = os.listdir(directory)
#         want_index_list = []
#         for i in file_name_list:
#             if file_group_name in i:
#                 want_index_list.append(i)
#
#         # 숫자만 축출
#         digit = []
#         for i in want_index_list:
#             digit_str = ''
#             for ii in i:
#                 if ii.isdigit():
#                     digit_str = digit_str + ii
#                 else:
#                     pass
#             digit.append(digit_str)
#
#         # 날짜에 대한 것만 축출
#         nowdate_list = []
#         for i in digit:
#             nowdate_list.append(i[-len_now_time:])
#
#         # 날짜를 정수로 변환
#         nowdate_list_int = []
#         for i in nowdate_list:
#             nowdate_list_int.append(int(i))
#
#         # 가장 큰 정수를 가진 항목의 위치를 구함
#         i = 0
#         for i in range(0, len(nowdate_list_int)):
#             if str(max(nowdate_list_int)) in nowdate_list[i]:
#                 break
#
#         return want_index_list[i]
#
#     @staticmethod
#     def get_window_list():
#         """
#         :return: 현재 띄워져있는 윈도우 리스트[list]
#         """
#
#         def callback(hwnd, hwnd_list: list):
#             title = win32gui.GetWindowText(hwnd)
#             if win32gui.IsWindowEnabled(hwnd) and win32gui.IsWindowVisible(hwnd) and title:
#                 hwnd_list.append((title, hwnd))
#             return True
#
#         output = []
#         win32gui.EnumWindows(callback, output)
#
#         return output
#
#     def run_focus_window(self, window_name):
#         """
#         :param window_name:
#         :return: 원하는 윈도우를 탑 윈도우로 하기(포커스 하기)[-]
#         """
#
#         # parameter
#         window_name = str(window_name)
#         upwindowlist = self.get_window_list()
#         hwnd = 0
#         for i in range(len(upwindowlist)):
#             if upwindowlist[i][0] == window_name:
#                 hwnd = upwindowlist[i][1]
#                 break
#
#         win32gui.ShowWindow(hwnd, 2)  # 2=> win32con.SW_SHOW
#         time.sleep(0.5)
#         win32gui.ShowWindow(hwnd, win32con.SW_MAXIMIZE)
#         time.sleep(0.5)
#         win32gui.ShowWindow(hwnd, win32con.SW_MAXIMIZE)
#         time.sleep(0.5)
#         win32gui.ShowWindow(hwnd, win32con.SW_MAXIMIZE)
#
#     def run_exit_window(self, window_name):
#         """
#         :param window_name:
#         :return: 원하는 윈도우 종료 하기[-]
#         """
#
#         # parameter
#         window_name = str(window_name)
#         upwindowlist = self.get_window_list()
#         hwnd = 0
#         for i in range(len(upwindowlist)):
#             if upwindowlist[i][0] == window_name:
#                 hwnd = upwindowlist[i][1]
#                 break
#         win32gui.PostMessage(hwnd, win32con.WM_CLOSE, 0, 0)
#
#     def get_array_of_special_application(self):
#         """
#         :return: monitor_board[numpy.array(6x2)]
#         """
#         # 행: (0, Helper), (1, POP DTS), (2, 삼성증권(에러 발생시)), (3, Microsoft Office 인증 마법사), (4, Excel)
#         # 정상적으로 POP DTS 팝업창이 뜰경우 POP DTS가 보이지 않음
#         # 열: (0, 꺼짐, 켜짐), (1, 켜진 개수)
#         monitor_board = numpy.zeros((6, 2), dtype=int)
#
#         win_list = self.get_window_list()
#         for i in win_list:
#             if i[0] == r'C:\WINDOWS\System32\cmd.exe':
#                 monitor_board[0][0] = 1
#                 monitor_board[0][1] += 1
#
#             elif i[0] == '삼성증권 POP DTS V1.0  [Family Center 1588-2323]':
#                 monitor_board[1][0] = 1
#                 monitor_board[1][1] += 1
#
#             elif i[0] == '삼성증권':
#                 monitor_board[2][0] = 1
#                 monitor_board[2][1] += 1
#
#             elif i[0] == 'Microsoft Office 인증 마법사':
#                 monitor_board[3][0] = 1
#                 monitor_board[3][1] += 1
#
#             elif 'Excel (제품 인증 실패)' in i[0]:
#                 monitor_board[4][0] = 1
#                 monitor_board[4][1] += 1
#
#             elif i[0] == 'ncStarter':
#                 monitor_board[5][0] = 1
#                 monitor_board[5][1] += 1
#
#         array_of_outside_application = monitor_board
#
#         return array_of_outside_application


class FileSystem:
    def __init__(self):
        pass

    @staticmethod
    def get_recent_file_name(directory, file_group_name, len_now_time, extension=None):
        # parameter
        directory = str(directory)
        file_group_name = str(file_group_name)
        len_now_time = int(len_now_time)

        file_name_list = os.listdir(directory)
        list_file_name = []
        for i in file_name_list:
            namee, ext = os.path.splitext(i)
            list_file_name.append(namee)

        want_index_list = []
        for i in list_file_name:
            if file_group_name in i:
                want_index_list.append(i)

        # 숫자만 축출
        digit = []
        for n in want_index_list:
            i = n[-len_now_time:]
            digit_str = ''
            for ii in i:
                if ii.isdigit():
                    digit_str = digit_str + ii
                else:
                    pass
            digit.append(digit_str)

        # 날짜에 대한 것만 축출
        nowdate_list = []
        for i in digit:
            nowdate_list.append(i[-len_now_time:])

        # 날짜를 정수로 변환
        nowdate_list_int = []
        for i in nowdate_list:
            nowdate_list_int.append(int(i))

        # 가장 큰 정수를 가진 항목의 위치를 구함
        i = 0
        for i in range(0, len(nowdate_list_int)):
            if str(max(nowdate_list_int)) in nowdate_list[i]:
                break

        if extension == None:
            anw = want_index_list[i]
        else:
            anw = want_index_list[i]

        return anw


class Memory:
    def __init__(self):
        self.ins_filesystem = FileSystem()

    def create_cover_json(self, directory, file_name, dic):
        path_file = "{}{}.json".format(directory, file_name)
        with open(path_file, 'w') as f:
            json.dump(dic, f)

    def create_cover_feather(self, directory, file_name, df):
        df = df.reset_index(drop=True)  # "index" 가 0 부터 1씩 아래로 늘어나도록 수정
        if len(df) == 0:
            df['0'] = numpy.nan

        pandas.options.display.float_format = '{:.4f}'.format
        df = df.astype(str)

        path_file = '{}{}.ftr'.format(directory, file_name)
        df.to_feather(path=path_file)

    def get_cover_saved_data(self, directory, file_name, extension=None):
        saved_data = None
        if extension == None:
            pass

        elif extension == 'json':
            path_file = "{}{}.json".format(directory, file_name)
            with open(path_file, 'r') as f:
                saved_data = json.load(f)

        elif extension == 'ftr':
            path_file = "{}{}.ftr".format(directory, file_name)
            time.sleep(0.2)
            saved_data = pandas.read_feather(path=path_file, columns=None, use_threads=True, storage_options=None)

        return saved_data

    def create_cumulate_feather(self, directory, file_name, df):
        df = df.reset_index(drop=True)  # "index" 가 0 부터 1씩 아래로 늘어나도록 수정
        if len(df) == 0:
            df['0'] = numpy.nan

        pandas.options.display.float_format = '{:.4f}'.format
        df = df.astype(str)

        now_time = datetime.datetime.now().strftime('%Y%m%d%H%M')
        file_name = '{}_{}'.format(file_name, now_time)
        path_file = '{}{}.ftr'.format(directory, file_name)
        df.to_feather(path=path_file)

    def create_cumulate_csv(self, directory, file_name, df):
        df = df.reset_index(drop=True)  # "index" 가 0 부터 1씩 아래로 늘어나도록 수정
        if len(df) == 0:
            df['0'] = numpy.nan

        pandas.options.display.float_format = '{:.4f}'.format
        df = df.astype(str)

        now_time = datetime.datetime.now().strftime('%Y%m%d%H%M')
        file_name = '{}_{}'.format(file_name, now_time)
        path_or_buf = '{}{}.csv'.format(directory, file_name)
        df.to_csv(path_or_buf=path_or_buf,
                  float_format='0.8f',
                  index=True,
                  index_label='index')

    def get_cumulate_recent_saved_data(self, len_now_time, directory, file_group_name, extension=None):
        saved_data = None

        file_name = self.ins_filesystem.get_recent_file_name(directory=directory,
                                                             file_group_name=file_group_name,
                                                             len_now_time=len_now_time,
                                                             extension=extension)
        filepath_or_buffer = '{}{}.{}'.format(directory, file_name, extension)
        if extension == 'csv':
            saved_data = pandas.read_csv(filepath_or_buffer=filepath_or_buffer,
                                         header=0,
                                         index_col='index',
                                         dtype={
                                             'minprice': str,
                                             'minqty': str,
                                             'minnotional': str
                                         },
                                         na_values='NaN',
                                         thousands=',')
        elif extension == 'ftr':
            saved_data = pandas.read_feather(path=filepath_or_buffer, columns=None, use_threads=True, storage_options=None)

        return saved_data


# class OutsideDevice:
#     @staticmethod
#     def run_click_mouse(xy, time_of_after_click):
#         pyautogui.click(xy)
#         time.sleep(time_of_after_click)
#
#     @staticmethod
#     def print_mouse_position():
#         while 1:
#             print(pyautogui.position())
#             time.sleep(2)
#
#     @staticmethod
#     def run_write_keyboard(text, period, time_of_after_press):
#         pyautogui.typewrite(text, period)
#         time.sleep(time_of_after_press)


class ETC:
    def __init__(self):
        self.ins_filesystem = FileSystem()

    def get_list_of_interest_stock(self):
        """
        :return: 종목코드, 종목명[list]
        """
        # excel file 읽기
        extension = 'xlsx'
        file_name = self.ins_filesystem.get_recent_file_name(directory='oasis_dir',
                                                           file_group_name='interest_stock_',
                                                           len_now_time=12,
                                                         extension=extension)
        wb = openpyxl.load_workbook('oasis_dir' + '/' + file_name + '.' + extension)
        sheet = wb.get_sheet_by_name('Sheet1')

        # 'excel'에서 column 알파벳 알아내기
        code_column_letter = 'A1'
        name_column_letter = 'A1'
        for column_num in range(1, sheet.max_column + 1):
            if sheet.cell(row=1, column=column_num).value == '종목코드':
                code_column_letter = get_column_letter(column_num)

            if sheet.cell(row=1, column=column_num).value == '종목명':
                name_column_letter = get_column_letter(column_num)

        # 투자 후보 종목명, 코드 리스트화
        stock_code_list = []
        candidate_stock_list = []

        for row_num in range(2, sheet.max_row + 1):
            stock_code = str(sheet[code_column_letter + str(row_num)].value)
            candidate_stock = str(sheet[name_column_letter + str(row_num)].value)

            stock_code_list.append(stock_code)
            candidate_stock_list.append(candidate_stock)

            print(stock_code + ': ' + candidate_stock)

        stock_list = candidate_stock_list

        return stock_code_list, stock_list

    @staticmethod
    def get_timestamp_to_datetime_date(timestamp):
        """
        :param timestamp:
        :return: [datetime.date]
        """

        datetime_date = timestamp.date()

        return datetime_date

    @staticmethod
    def get_trans_price(price):
        """
        :param price:
        :return: 다음 장 가격 예측 값을 실거래가에 맞게 변환[int]
        """

        price = int(price)

        if price < 1000:
            trans_price = price

        elif 1000 <= price < 5000:
            price = str(price)
            price = list(price)
            if int(price[3]) > 7 or int(price[3]) < 3:
                price = ''.join(price)
                trans_price = round(int(price), -1)
            elif 3 <=int(price[3]) < 5 or 5 < int(price[3]) <= 7:
                price[3] = '5'
                price = ''.join(price)
                trans_price = int(price)
            else:
                price = ''.join(price)
                trans_price = int(price)

        elif 5000 <= price < 10000:
            trans_price = round(price, -1)

        elif 10000 <= price < 50000:
            price = str(price)
            if int(price[-2:]) > 75 or int(price[-2:]) < 25:
                trans_price = round(int(price), -2)
            elif 25 <=int(price[-2:]) < 50 or 50 < int(price[-2:]) <= 75:
                price = list(price)
                price[-2] = '5'
                price[-1] = '0'
                price = ''.join(price)
                trans_price = int(price)
            else:
                trans_price = int(price)

        elif 50000 <= price < 100000:
            trans_price = round(price, -2)

        elif 100000 <= price < 500000:
            price = str(price)
            if int(price[-3:]) > 750 or int(price[-3:]) < 250:
                trans_price = round(int(price), -3)
            elif 250 <=int(price[-3:]) < 500 or 500 < int(price[-3:]) <= 750:
                price = list(price)
                price[-3] = '5'
                price[-2] = '0'
                price[-1] = '0'
                price = ''.join(price)
                trans_price = int(price)
            else:
                trans_price = int(price)

        elif 500000 <= price:
            trans_price = round(price, -3)

        else:
            trans_price = 'error'

        return trans_price

    # 시간을 고려하지 않은 최악의 함수다.
    # def get_minute_caled(self, min_init, min_variance):
    #     minute_cald = min_init + min_variance
    #     if 0 < minute_cald:
    #         minute_cald = minute_cald % 60
    #     elif 0 == minute_cald:
    #         minute_cald = 0
    #     else:
    #         minute_cald = abs(minute_cald) % 60
    #         minute_cald = 60 - minute_cald
    #
    #     return minute_cald

    # sys.path.extend(['/home/woong/20210205_project/20210205_code/20210821_insight/composition']) 경로추가
    # os.chdir(r'C:\imjiwoong\20200331_with_python\20200331_oasis\20200704_alpha3\auto_trading')  ???

    @staticmethod
    def get_min_win_rate(profit_rate, fee):
        x = Symbol('x')
        y = x*(profit_rate-fee*(2+profit_rate)) - (100-x)*(profit_rate+(fee*(2+profit_rate)))
        x_anw = solve(y)

        min_win_rate = x_anw

        return min_win_rate

    def save_load(self):
        # df
        df = 0
        # 중복 데이터 제거
        df = df.drop_duplicates(['time'], keep='last')

        # 일자 기준 오름차순으로 변경
        df = df.sort_values(by='time')

        # index 가 0 부터 1씩 아래로 늘어나도록 수정
        df = df.reset_index(drop=True)

        # csv
        df = 0
        df = df.astype(str)
        file_name = ''
        now_time = ''
        extension = ''
        path_or_buf = '{}/{}_{}.csv'.format(self.ins_env_directory.log, file_name, now_time)
        df.to_csv(path_or_buf=path_or_buf,
                  index=True,
                  index_label='index')

        filepath_or_buffer = '{}/{}.{}'.format(self.ins_env_directory.binance, file_name, extension)
        df = pandas.read_csv(filepath_or_buffer=filepath_or_buffer,
                             header=0,
                             index_col='index',
                             dtype={
                                 'minprice': str,
                                 'minqty': str,
                                 'minnotional': str
                             },
                             na_values='NaN',
                             thousands=',')

        # feather
        df = df.astype(str)
        file_name = 'log_trade'
        path = './{}.ftr'.format(file_name)
        df.to_feather(path=path)

        file_name = 'log_trade'
        path = './{}.ftr'.format(file_name)
        df = pandas.read_feather(path=path, columns=None, use_threads=True, storage_options=None)

        # npy
        cleanup_data_std = 0
        arr = numpy.array(cleanup_data_std)
        file_name = 'cleanup_data_std'
        extension = 'npy'
        file = '{}/{}.{}'.format(self.ins_env_directory.feed, file_name, extension)
        numpy.save(file=file, arr=arr)

        file_name = 'input_of_shell'
        extension = 'npy'
        file = '{}/{}.{}'.format(self.ins_env_directory.feed, file_name, extension)
        input_of_shell = numpy.load(file=file)


class TimeKJW:
    def __init__(self):
        pass

    def Contorol(self):
        tm = 1234567890
        tm = datetime.datetime.fromtimestamp(tm)
        t2 = tm.replace(minute=00)
        t2 = int(time.mktime(t2.timetuple()))

        tm = time.localtime(tm)
        time_minute = tm.tm_min


class DF:
    def __init__(self):
        # price is column
        index_max = price.idxmax()
        self.term_high_price = price.loc[index_max]


if __name__ == '__main__':
    # ins_ubuntu = Ubuntu()
    # ins = SpecialApplication()
    # aa = ins.pop_dts_wn
    # a = ins_ubuntu.get_recent_file_name(
    #     directory='/home/woong/20210205_project/20210205_data/20210605_binance',
    #     file_group_name='minprice_minqty_minnotional_',
    #     len_now_time=12,
    #     extension='.csv'
    # )
    ins = ETC()
    # min_win_rate = ins.get_min_win_rate(0.045, 0.0004)  # 51
    # min_win_rate = ins.get_min_win_rate(0.0032, 0.0004)  # 62.52, 내 승률은 58.68
    # min_win_rate = ins.get_min_win_rate(0.0081, 0.0004)  # 54.96
    min_win_rate = ins.get_min_win_rate(0.0101, 0.0004)  # 53.98

    # for i in range(1, 101):
    #     print(i*49.58 - ((100-i)*50.42))
